import csv
import io
from typing import List, Optional
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl

from app.auth import User, require_role
from app.services.shared import question_bank_service

router = APIRouter(prefix="/question-bank", tags=["question-bank"])
service = question_bank_service


class UpdateQuestionPayload(BaseModel):
    category: Optional[str] = None
    topic: Optional[str] = None
    difficulty: Optional[str] = None
    question_text: Optional[str] = None
    expected_answer: Optional[str] = None
    keywords: Optional[List[str]] = None



@router.post("/import")
def import_questions(payload: list[dict], user: User = Depends(require_role("admin", "trainer"))) -> dict:
    imported = service.import_questions(payload)
    return {"count": len(imported), "questions": [q.__dict__ for q in imported]}


@router.get("/search")
def search_questions(query: str, user: User = Depends(require_role("learner", "admin", "trainer"))) -> list[dict]:
    return [q.__dict__ for q in service.search(query)]


@router.get("/all")
def list_questions(user: User = Depends(require_role("learner", "admin", "trainer"))) -> list[dict]:
    return [q.__dict__ for q in service.list_questions()]


@router.post("/bulk")
def bulk_import_questions(payload: list[dict], user: User = Depends(require_role("admin", "trainer"))) -> dict:
    imported = service.import_questions(payload)
    return {"count": len(imported), "questions": [q.__dict__ for q in imported]}


@router.post("/upload")
def upload_question_bank(file: UploadFile = File(...), user: User = Depends(require_role("admin", "trainer"))) -> dict:
    filename = file.filename or ""
    questions: List[dict] = []

    try:
        content = file.file.read()
        
        if filename.endswith(".csv"):
            text = content.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text))
            
            # Normalize headers
            headers = [h.strip().lower() for h in (reader.fieldnames or [])]
            for row in reader:
                # Map headers dynamically
                row_data = {k.strip().lower(): v for k, v in row.items() if k}
                keywords_raw = row_data.get("keywords", "")
                keywords = [kw.strip() for kw in keywords_raw.replace(",", "|").replace(";", "|").split("|") if kw.strip()]
                
                questions.append({
                    "category": row_data.get("category", "general"),
                    "topic": row_data.get("topic", "general"),
                    "difficulty": row_data.get("difficulty", "medium"),
                    "question_text": row_data.get("question_text", row_data.get("question", "")),
                    "expected_answer": row_data.get("expected_answer", row_data.get("answer", "")),
                    "keywords": keywords,
                    "source": filename
                })
                
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            if not sheet:
                raise HTTPException(status_code=400, detail="Empty sheet in Excel file")
                
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="No rows found in Excel sheet")
                
            headers = [str(cell).strip().lower() for cell in rows[0] if cell is not None]
            
            for row in rows[1:]:
                if not any(row):  # skip empty rows
                    continue
                row_dict = {}
                for idx, cell_value in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = cell_value
                
                keywords_raw = str(row_dict.get("keywords", "")) if row_dict.get("keywords") is not None else ""
                keywords = [kw.strip() for kw in keywords_raw.replace(",", "|").replace(";", "|").split("|") if kw.strip()]
                
                questions.append({
                    "category": str(row_dict.get("category", "general")),
                    "topic": str(row_dict.get("topic", "general")),
                    "difficulty": str(row_dict.get("difficulty", "medium")),
                    "question_text": str(row_dict.get("question_text", row_dict.get("question", ""))),
                    "expected_answer": str(row_dict.get("expected_answer", row_dict.get("answer", ""))),
                    "keywords": keywords,
                    "source": filename
                })
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload CSV or XLSX.")
            
        if not questions:
            raise HTTPException(status_code=400, detail="No questions parsed from the file.")
            
        imported = service.import_questions(questions)
        return {"count": len(imported), "questions": [q.__dict__ for q in imported]}
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


@router.get("/export")
def export_question_bank(user: User = Depends(require_role("admin", "trainer"))):
    questions = service.list_questions()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    
    headers = ["category", "topic", "difficulty", "question_text", "expected_answer", "keywords"]
    ws.append(headers)
    
    for q in questions:
        ws.append([
            q.category,
            q.topic,
            q.difficulty,
            q.question_text,
            q.expected_answer,
            ", ".join(q.keywords)
        ])
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=questions_export.xlsx"}
    )


@router.put("/{id}")
def update_question(id: int, payload: UpdateQuestionPayload, user: User = Depends(require_role("admin", "trainer"))) -> dict:
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updated = service.update_question(id, updates)
    if not updated:
        raise HTTPException(status_code=404, detail="Question not found")
    return updated.__dict__


@router.delete("/{id}")
def delete_question(id: int, user: User = Depends(require_role("admin", "trainer"))) -> dict:
    success = service.delete_question(id)
    if not success:
        raise HTTPException(status_code=404, detail="Question not found")
    return {"status": "success", "message": f"Question {id} deleted successfully"}


